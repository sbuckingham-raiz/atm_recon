import os
import datetime

import openpyxl
import openpyxl.workbook
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from common.utility import get_directory_of_file

file_dir = get_directory_of_file(__file__)

TEMPLATE_FILE = os.path.join(file_dir, 'recon_file_template.xlsx')

def write_data(total_df:pd.DataFrame, device_amts_df:pd.DataFrame, start_end_amts:pd.DataFrame, workbook:openpyxl.workbook.workbook.Workbook, current_date:datetime.datetime=None):
    if current_date == None:
        current_date = datetime.datetime.today()
        

    orangeColor = 'FFF4B084'
    orangeFill = PatternFill(start_color=orangeColor, end_color=orangeColor, fill_type='solid')
    
    lightBlueColor = 'FF91D2FF'
    lightBlueFill = PatternFill(start_color=lightBlueColor, end_color=lightBlueColor, fill_type='solid')
    
    darkBlueColor = 'FF5B9BD5'
    darkBlueFill = PatternFill(start_color=darkBlueColor, end_color=darkBlueColor, fill_type='solid')

    sheetNames = workbook.sheetnames
    for sheetName in sheetNames:
        try:
            sheet = workbook[sheetName]
            freezeCell = sheet['B2']
            sheet.freeze_panes = freezeCell
            sheet_total_df = total_df[total_df['device_name']==sheetName]
            sheet_device_amts_df = device_amts_df[device_amts_df['device_name']==sheetName]
            sheet_start_end_amts = start_end_amts[start_end_amts['device_name']==sheetName]


            startRow = 1

            currentCell = sheet.cell(row=startRow, column=1).value
            currentRow = startRow


            while currentCell:
                currentRow += 1
                currentCell = sheet.cell(row=currentRow, column=1).value

                if not currentCell:
                    break
                
                secondColCell = sheet.cell(row=currentRow, column=2).value

                currentCellDate = datetime.datetime.strptime(currentCell, '%Y-%m-%d')

                if currentCellDate > current_date:
                    break
                
                if secondColCell:
                    continue
                
                cell_total_df = sheet_total_df[sheet_total_df['date']==currentCellDate]              
                cell_start_end_amts = sheet_start_end_amts[sheet_start_end_amts['date']==currentCellDate]
                start_end_amts_checks = cell_start_end_amts[cell_start_end_amts['device_type']=='Check Scanner']
                cell_device_amts_df  = sheet_device_amts_df[sheet_device_amts_df['date']==currentCellDate]

                ## ------------------- Col B --------------------------------------------------------------------------------------------------------------------------- 
                beginning_amount_values = cell_total_df['previous_amount'].dropna().values
                beginningAmountCell = sheet.cell(row=currentRow, column=2)
                beginningAmountCell.fill = orangeFill
                if beginning_amount_values.any():
                    beginning_amount = beginning_amount_values[0]
                    beginningAmountCell.value = beginning_amount
                    beginningAmountCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


                ## ------------------- Col C --------------------------------------------------------------------------------------------------------------------------- 
                cash_acceptor_df_banker = cell_device_amts_df[(cell_device_amts_df['device_type']=='Cash Acceptor') & (cell_device_amts_df['post_type']=='Banker')]
                cash_acceptor_df_banker_values = cash_acceptor_df_banker['amount_increased'].dropna().values
                if cash_acceptor_df_banker_values.any():
                    cashAcceptorBankerCashInCell = sheet.cell(row=currentRow, column=3)
                    cash_acceptor_banker_cash_in = cash_acceptor_df_banker_values[0]
                    cashAcceptorBankerCashInCell.value = cash_acceptor_banker_cash_in
                    cashAcceptorBankerCashInCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col D ---------------------------------------------------------------------------------------------------------------------------
                cash_acceptor_df_atm = cell_device_amts_df[(cell_device_amts_df['device_type']=='Cash Acceptor') & (cell_device_amts_df['post_type']=='ATM')]
                cash_acceptor_df_atm_values = cash_acceptor_df_atm['amount_increased'].dropna().values

                if cash_acceptor_df_atm_values.any():
                    cashAcceptorATMCashInCell = sheet.cell(row=currentRow, column=4)
                    cash_acceptor_atm_cash_in = cash_acceptor_df_atm_values[0]
                    cashAcceptorATMCashInCell.value = cash_acceptor_atm_cash_in
                    cashAcceptorATMCashInCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  

                ## ------------------- Col E ---------------------------------------------------------------------------------------------------------------------------
                cash_acceptor_df_CIT = cell_device_amts_df[(cell_device_amts_df['device_type']=='Cash Acceptor') & (cell_device_amts_df['post_type']=='CIT')]
                cash_acceptor_df_CIT_values = cash_acceptor_df_CIT['amount_decreased'].dropna().values

                if cash_acceptor_df_CIT_values.any():
                    cash_acceptor_CIT_cash_removed = cash_acceptor_df_CIT_values[0]
                    cashAcceptorCITCashInCell = sheet.cell(row=currentRow, column=5)
                    cashAcceptorCITCashInCell.value = cash_acceptor_CIT_cash_removed
                    cashAcceptorCITCashInCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col F ---------------------------------------------------------------------------------------------------------------------------
                cash_dispenser_df_CIT = cell_device_amts_df[(cell_device_amts_df['device_type']=='Cash Dispenser') & (cell_device_amts_df['post_type']=='CIT')]
                cash_dispenser_df_CIT_values = cash_dispenser_df_CIT['amount_increased'].dropna().values

                if cash_dispenser_df_CIT_values.any():
                    cashDispenserCITCashRemovedCell = sheet.cell(row=currentRow, column=6)
                    cash_dispenser_CIT_cash_in = cash_dispenser_df_CIT_values[0]
                    cashDispenserCITCashRemovedCell.value = cash_dispenser_CIT_cash_in
                    cashDispenserCITCashRemovedCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col G ---------------------------------------------------------------------------------------------------------------------------
                cash_dispenser_banker_df = cell_device_amts_df[(cell_device_amts_df['device_type']=='Cash Dispenser') & (cell_device_amts_df['post_type']=='Banker')]
                cash_dispenser_banker_df_values_cash_out = cash_dispenser_banker_df['amount_decreased'].dropna().values

                if cash_dispenser_banker_df_values_cash_out.any():
                    cashDispenserBankerCashRemovedCell = sheet.cell(row=currentRow, column=7)
                    cash_dispenser_Banker_cash_out = cash_dispenser_banker_df_values_cash_out[0]
                    cashDispenserBankerCashRemovedCell.value = cash_dispenser_Banker_cash_out
                    cashDispenserBankerCashRemovedCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col H ---------------------------------------------------------------------------------------------------------------------------
                cash_dispenser_ATM_df = cell_device_amts_df[(cell_device_amts_df['device_type']=='Cash Dispenser') & (cell_device_amts_df['post_type']=='ATM')]
                cash_dispenser_ATM_df_values_cash_out = cash_dispenser_ATM_df['amount_decreased'].dropna().values

                if cash_dispenser_ATM_df_values_cash_out.any():
                    cash_dispenser_ATM_cash_out = cash_dispenser_ATM_df_values_cash_out[0]
                    cashDispenserATMCashRemovedCell = sheet.cell(row=currentRow, column=8)
                    cashDispenserATMCashRemovedCell.value = cash_dispenser_ATM_cash_out
                    cashDispenserATMCashRemovedCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'                


                ## ------------------- Col I --------------------------------------------------------------------------------------------------------------------------
                cash_dispenser_CIT_df = cell_device_amts_df[(cell_device_amts_df['device_type']=='Cash Dispenser') & (cell_device_amts_df['post_type']=='CIT')]
                cash_dispenser_CIT_df_values_cash_out = cash_dispenser_CIT_df['amount_decreased'].dropna().values

                if cash_dispenser_CIT_df_values_cash_out.any():
                    cash_dispenser_CIT_cash_out = cash_dispenser_CIT_df_values_cash_out[0]
                    cashDispenserCITCashRemovedCell = sheet.cell(row=currentRow, column=9)
                    cashDispenserCITCashRemovedCell.value = cash_dispenser_CIT_cash_out
                    cashDispenserCITCashRemovedCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)' 

                ## ------------------- Col J and K --------------------------------------------------------------------------------------------------------------------------
                coin_dispenser_CIT_df = cell_device_amts_df[(cell_device_amts_df['device_type']=='Coin Dispenser') & (cell_device_amts_df['post_type']=='CIT')]
                coin_dispenser_CIT_df_values_cash_in = coin_dispenser_CIT_df['amount_increased'].dropna().values
                coin_dispenser_CIT_df_values_cash_out = coin_dispenser_CIT_df['amount_decreased'].dropna().values

                if coin_dispenser_CIT_df_values_cash_in.any():
                    coin_dispenser_CIT_cash_in = coin_dispenser_CIT_df_values_cash_in[0]
                    coinDispenserCITCashInCell = sheet.cell(row=currentRow, column=10)
                    coinDispenserCITCashInCell.value = coin_dispenser_CIT_cash_in
                    coinDispenserCITCashInCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)' 

                if coin_dispenser_CIT_df_values_cash_out.any():
                    coin_dispenser_CIT_cash_out = coin_dispenser_CIT_df_values_cash_out[0]
                    coinDispenserCITCashOutCell = sheet.cell(row=currentRow, column=11)
                    coinDispenserCITCashOutCell.value = coin_dispenser_CIT_cash_out
                    coinDispenserCITCashOutCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col N ---------------------------------------------------------------------------------------------------------------------------
                ending_amount_values = cell_total_df['amount'].dropna().values
                endingAmountCell = sheet.cell(row=currentRow, column=14)
                endingAmountCell.fill = orangeFill

                if ending_amount_values.any():
                    ending_amount = ending_amount_values[0]
                    endingAmountCell.value = ending_amount
                    endingAmountCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col O ---------------------------------------------------------------------------------------------------------------------------
                estimated_amount_values = cell_total_df['estimated_amount'].dropna().values
                ColOformula = f'=B{currentRow} + C{currentRow} + D{currentRow} - E{currentRow} + F{currentRow} - G{currentRow} - H{currentRow} - I{currentRow} + J{currentRow} - K{currentRow} + L{currentRow} - M{currentRow}'

                if estimated_amount_values.any():
                    estimatedAmountCell = sheet.cell(row=currentRow, column=15)
                    estimatedAmountCell.value = ColOformula
                    estimatedAmountCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


                ## ------------------- Col P ---------------------------------------------------------------------------------------------------------------------------
                delta_values = cell_total_df['delta'].dropna().values
                ColPformula = f'=N{currentRow} - O{currentRow}'
                deltaAmountCell = sheet.cell(row=currentRow, column=16)
                deltaAmountCell.fill = lightBlueFill

                deltaAmountCell.value = ColPformula
                deltaAmountCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col S ---------------------------------------------------------------------------------------------------------------------------
                ColSformula = f'=N{currentRow} - Q{currentRow} - R{currentRow}'
                ColSCell = sheet.cell(row=currentRow, column=19)
                ColSCell.value = ColSformula
                ColSCell.fill = orangeFill
                ColSCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col Z (26) ---------------------------------------------------------------------------------------------------------------------------
                ColZformula = f'=S{currentRow} - T{currentRow} + U{currentRow} - V{currentRow}'
                ColZCell = sheet.cell(row=currentRow, column=26)
                ColZCell.value = ColZformula
                ColZCell.fill = orangeFill
                ColZCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col AB (28) ---------------------------------------------------------------------------------------------------------------------------
                check_amt_starting_values = start_end_amts_checks['starting_amt'].dropna().values
                checkStartingAmountCell = sheet.cell(row=currentRow, column=28)
                checkStartingAmountCell.fill = darkBlueFill

                if check_amt_starting_values.any():
                    check_starting_amt = check_amt_starting_values[0]
                    checkStartingAmountCell.value = check_starting_amt
                    checkStartingAmountCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col AC (29) ---------------------------------------------------------------------------------------------------------------------------
                check_scanner_banker_df = cell_device_amts_df[(cell_device_amts_df['device_type']=='Check Scanner') & (cell_device_amts_df['post_type']=='Banker')]
                check_scanner_banker_df_values_cash_in = check_scanner_banker_df['amount_increased'].dropna().values

                if check_scanner_banker_df_values_cash_in.any():
                    check_scanner_banker_cash_in = check_scanner_banker_df_values_cash_in[0]
                    checkScannerBankerCashInCell = sheet.cell(row=currentRow, column=29)
                    checkScannerBankerCashInCell.value = check_scanner_banker_cash_in
                    checkScannerBankerCashInCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col AD (30) ---------------------------------------------------------------------------------------------------------------------------
                check_scanner_CIT_df = cell_device_amts_df[(cell_device_amts_df['device_type']=='Check Scanner') & (cell_device_amts_df['post_type']=='CIT')]
                check_scanner_CIT_df_values_cash_out = check_scanner_CIT_df['amount_decreased'].dropna().values

                if check_scanner_CIT_df_values_cash_out.any():
                    check_scanner_CIT_cash_out = check_scanner_CIT_df_values_cash_out[0]
                    checkScannerCITCashOutCell = sheet.cell(row=currentRow, column=30)
                    checkScannerCITCashOutCell.value = check_scanner_CIT_cash_out
                    checkScannerCITCashOutCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col AE (31) ---------------------------------------------------------------------------------------------------------------------------
                check_amt_ending_values = start_end_amts_checks['ending_amt'].dropna().values
                checkEndingAmountCell = sheet.cell(row=currentRow, column=31)
                checkEndingAmountCell.fill = darkBlueFill

                if check_amt_ending_values.any():
                    check_ending_amt = check_amt_ending_values[0]
                    checkEndingAmountCell.value = check_ending_amt
                    checkEndingAmountCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col AF (32) ---------------------------------------------------------------------------------------------------------------------------
                ColAFformula = f'=AB{currentRow} + AC{currentRow} - AD{currentRow}'

                ColAFCell = sheet.cell(row=currentRow, column=32)
                ColAFCell.value = ColAFformula
                ColAFCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                ## ------------------- Col AG (33) ---------------------------------------------------------------------------------------------------------------------------
                ColAGformula = f'=AE{currentRow} - AF{currentRow}'

                ColAGCell = sheet.cell(row=currentRow, column=33)
                ColAGCell.value = ColAGformula
                ColAGCell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

            autofit_columns(sheet)
        except Exception as e:
            print(f'Error running for sheet name {sheetName}. {e}')
        
    return workbook
    
def create_new_template(pick_date:datetime.datetime, device_names:list, dates_after:datetime):
    workbook = openpyxl.load_workbook(TEMPLATE_FILE)
    sheet = workbook.active
    first_date = sheet.cell(row=2, column=1).value
    
    if not first_date:
        current_year = pick_date.year
        start_date = datetime.date(year=current_year, day=1, month=1)
        end_date = datetime.date(year=current_year, month=12, day=31)

        current_date = start_date
        current_row = 2
        while current_date <= end_date:
            
            if current_date >= dates_after:
                sheet.cell(current_row, 1, value=current_date.strftime('%Y-%m-%d'))
                current_date =  current_date + datetime.timedelta(1)
                current_row+=1
            
    sheet.title = device_names[0]
    x = 0
    for device in device_names[1:]:
        workbook.copy_worksheet(workbook.active)
        x+=1
        s = workbook[workbook.sheetnames[x]]
        s.title = device
        
    return workbook

def autofit_columns(sheet):
    row_height = 60
    sheet.row_dimensions[0].height = row_height
    
    for column_cells in sheet.columns:
        col_letter = get_column_letter(column_cells[0].column)
        max_length = max([len(str(cell.value) or "") for cell in column_cells])
        max_length_margins = .8
        
        if col_letter in ['A', 'P']:
            max_length_margins = .75
            
        
        max_length = (max_length + 2) * max_length_margins
        
                
        sheet.column_dimensions[col_letter].width = max_length
        